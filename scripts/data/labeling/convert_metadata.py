#!/usr/bin/env python3
"""
Convert Student-Annotated Metadata to Machine-Readable Format

This script:
1. Loads raw metadata.xlsx (human-annotated by students)
2. Converts Chinese column names to English
3. Standardizes the format and preserves or generates filenames
4. Saves as labels.csv for downstream processing

Note: This does NOT perform labeling - students have already labeled the data.
This script only converts Excel → CSV format.

Usage:
    python -m scripts.data.labeling.convert_metadata
"""

import csv
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

# Resolve paths relative to the validated repository root.
from diatom_cascade.config.path_config import get_data_root, get_project_root

PROJECT_ROOT = get_project_root()
DATA_ROOT = get_data_root(PROJECT_ROOT)
XLSX = DATA_ROOT / "raw" / "metadata.xlsx"
OUT = DATA_ROOT / "raw" / "labels.csv"
OUT.parent.mkdir(parents=True, exist_ok=True)

# Column name mapping
colmap = {
    '纲': 'class',
    '目': 'order',
    '科': 'family',
    '属名': 'genus',
    '种名': 'species',
    'ppt序号': 'ppt_index',  # actual column name in metadata.xlsx
    '文件名': 'filename',
}

wb = load_workbook(XLSX, data_only=True)
ws = wb.active

# Read headers
headers = [c.value if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
idx = {colmap.get(h, h): i for i, h in enumerate(headers)}

required_columns = {"class", "order", "family", "genus", "species"}
missing_columns = sorted(required_columns - set(idx))
if missing_columns:
    raise ValueError(f"metadata.xlsx is missing columns: {missing_columns}")
if "filename" not in idx and "ppt_index" not in idx:
    raise ValueError("metadata.xlsx requires either 'filename' or 'ppt序号'")


def get(row, key, default=""):
    i = idx.get(key)
    if i is None:
        return default
    v = row[i].value
    return "" if v is None else str(v).strip()


rows_out = []
for row in ws.iter_rows(min_row=2):  # Skip header row
    filename = get(row, 'filename')
    if not filename:
        ppt = get(row, 'ppt_index')
        if not ppt:
            continue
        try:
            ppt_i = int(float(ppt))
        except (TypeError, ValueError):
            print(f"Skipping worksheet row {row[0].row}: invalid ppt_index={ppt!r}")
            continue
        filename = f"slide_{ppt_i:04d}.png"
    genus = get(row, 'genus')
    species = get(row, 'species')

    rows_out.append({
        "filename": filename,
        "class": get(row, 'class'),
        "order": get(row, 'order'),
        "family": get(row, 'family'),
        "genus": genus,
        "species": species
    })

filenames = [row["filename"] for row in rows_out]
if not filenames:
    raise ValueError("metadata.xlsx contains no usable data rows")
duplicates = sorted(name for name, count in Counter(filenames).items() if count > 1)
if duplicates:
    raise ValueError(f"metadata.xlsx contains duplicate filenames: {duplicates[:10]}")

with OUT.open("w", newline="", encoding="utf-8-sig") as f:
    w = csv.DictWriter(f, fieldnames=["filename", "class", "order", "family", "genus", "species"])
    w.writeheader()
    w.writerows(rows_out)

print(f"Saved: {OUT} rows={len(rows_out)}")


